import itertools
import math
import numpy
from typing import List
import openpyxl as xl
import openpyxl.chart.bar_chart as barChart
import openpyxl.chart.line_chart as lineChart
import openpyxl.chart.reference as reference
import openpyxl.chart.series as series



class randomnessTests():
    def __init__(self, data_array: List[float]):
        self.data = data_array
        if(data_array == None or len(data_array) < 2):
            raise Exception('Array too short. Please use a bigger array')

    def __accumulate_array(self, array):
        cumulative_array = []
        cumulative_array.append(array[0])

        for i in range(1, len(array)):
            cumulative_array.append(array[i] + cumulative_array[i-1])
        return cumulative_array

    def __kolmogorov_smirnov(self, observed_frequency: List[int], expected_probability: List[float]):

        observed_frequency_total = sum(observed_frequency)
        observed_probability = [f / observed_frequency_total for f in observed_frequency]

        observed_cumulative_probability = self.__accumulate_array(observed_probability)
        expected_cumulative_probability = self.__accumulate_array(expected_probability)

        kolmogorov_smirnov_5percent = 1.63/numpy.sqrt(sum(observed_frequency))
        kolmogorov_smirnov_calculated = max([abs(expected_cumulative_probability[i] - observed_cumulative_probability[i]) for i in range(len(observed_cumulative_probability))])

        return (kolmogorov_smirnov_calculated < kolmogorov_smirnov_5percent)


    def uniformity_test(self, precision: int = 1, spreadsheet: bool = False, spreadsheet_title: str = 'uniformidade'):
        data_array = self.data.copy()

        if(precision < 1): precision = 1
        if(precision > 1000): precision = 1000

        num_classes = 10 * precision
        expected_probability = [1/num_classes for i in range(num_classes)]
        observed_frequency = numpy.histogram(a=data_array, range=[0,1], bins=num_classes)[0]
        expected_frequency = [n * len(data_array) for n in expected_probability]
        classes = self.__accumulate_array(expected_probability)

        #excel
        if(spreadsheet):
            wb = xl.Workbook()
            sheet = wb['Sheet']

            sheet.__setattr__('title', 'Corridas Ascendente')

            sheet.append(['classes', 'Frequência Observada', 'Frequência Esperada'])
            sheet.column_dimensions['B'].width = 21
            sheet.column_dimensions['C'].width = 19

            for i in range(num_classes):
                sheet.cell(row = (i+2), column = 1, value=classes[i])
                sheet.cell(row = (i+2), column = 2, value=observed_frequency[i])
                sheet.cell(row = (i+2), column = 3, value=expected_frequency[i])

            chart1 = lineChart.LineChart()
            chart1.title = 'Teste de Uniformidade'
            chart1.y_axis.title = 'Frequência'
            chart1.x_axis.title = 'Classes'

            graph_row = str(len(classes) + 5)

            data = reference.Reference(sheet, min_col=2, min_row=2, max_row=len(classes)+1, max_col=3)
            cats = reference.Reference(sheet, min_col=1, min_row=2, max_row=len(classes)+1)
            chart1.add_data(data)
            chart1.set_categories(cats)
            chart1.shape = 4
            chart1.series[0].title = series.SeriesLabel(v='Frequência Observada')
            chart1.series[1].title = series.SeriesLabel(v='Frequência Esperada')
            chart1.width = 30
            chart1.height = 15
            sheet.add_chart(chart1, 'A' + graph_row)

            wb.save(spreadsheet_title + '.xlsx')   



        return self.__kolmogorov_smirnov(observed_frequency, expected_probability)


    def runs_test(self, spreadsheet: bool = False, spreadsheet_title: str = 'corridas'):
        data_array = self.data.copy()

        def compute_runs(data_array, ascending = True):
            if(len(data_array) < 2): return
            
            runs = []
            run_count = 1
            index = 1

            while True:
                if( (ascending and data_array[index] > data_array[index-1]) or (not ascending and data_array[index] < data_array[index-1]) ): 
                    run_count = run_count + 1

                else:
                    runs.append(run_count)
                    run_count = 1
                    index = index + 1
                    
                index = index + 1
                if(index >= len(data_array)):
                    if(run_count > 1): runs.append(run_count)
                    break

            return runs

        ascending_runs = compute_runs(data_array)
        descending_runs = compute_runs(data_array, False)

        num_ascending_runs = [n for n in range(1, max(ascending_runs) + 1)]
        num_descending_runs = [n for n in range(1, max(descending_runs) + 1)]

        observed_frequency_ascending = [ascending_runs.count(n) for n in num_ascending_runs]
        observed_frequency_descending = [descending_runs.count(n) for n in num_descending_runs]

        expected_probability_ascending = [n/math.factorial(n + 1) for n in num_ascending_runs]
        expected_probability_descending = [n/math.factorial(n + 1) for n in num_descending_runs]

        
        #grafico
        expected_frequency_ascending = [round(n*sum(observed_frequency_ascending)) for n in expected_probability_ascending]
        expected_frequency_descending = [round(n*sum(observed_frequency_descending)) for n in expected_probability_descending]


        def data_to_excel(title: str, ascending: bool, sheet):
            sheet.append(['classes', 'Frequência Observada', 'Frequência Esperada'])
            sheet.column_dimensions['B'].width = 21
            sheet.column_dimensions['C'].width = 19

            if(ascending):
                classes = num_ascending_runs
                obs_freq = observed_frequency_ascending
                expected_freq = expected_frequency_ascending
            
            else:
                classes = num_descending_runs
                obs_freq = observed_frequency_descending
                expected_freq = expected_frequency_descending

            for i in range(len(classes)):
                sheet.cell(row = (i+2), column = 1, value=classes[i])
                sheet.cell(row = (i+2), column = 2, value=obs_freq[i])
                sheet.cell(row = (i+2), column = 3, value=expected_freq[i])

            chart1 = barChart.BarChart()
            chart1.type = "col"
            chart1.style = 10 
            chart1.title = title
            chart1.y_axis.title = 'Frequência'
            chart1.x_axis.title = 'Classes'


            graph_row = str(len(classes) + 5)

            data = reference.Reference(sheet, min_col=2, min_row=2, max_row=len(classes)+1, max_col=3)
            cats = reference.Reference(sheet, min_col=1, min_row=2, max_row=len(classes)+1)
            chart1.add_data(data)
            chart1.set_categories(cats)
            chart1.shape = 4
            chart1.series[0].title = series.SeriesLabel(v='Frequência Observada')
            chart1.series[1].title = series.SeriesLabel(v='Frequência Esperada')
            chart1.width = 30
            chart1.height = 15
            sheet.add_chart(chart1, 'A' + graph_row)


        if(spreadsheet):
            wb = xl.Workbook()
            sheet = wb['Sheet']

            sheet.__setattr__('title', 'Corridas Ascendente')

            data_to_excel('Corridas Ascendente', True, sheet)
            data_to_excel('Corridas Descendente', False, wb.create_sheet('Corridas Descendente'))

            wb.save(spreadsheet_title + '.xlsx')


        
        return self.__kolmogorov_smirnov(observed_frequency_ascending, expected_probability_ascending) and self.__kolmogorov_smirnov(observed_frequency_descending, expected_probability_descending)


    def gap_test(self, num_decimal_places: int = 1, digit: int = 0, spreadsheet: bool = False, spreadsheet_title: str = 'intervalos'):
        data_array = self.data.copy()

        if(num_decimal_places < 1 or num_decimal_places == None): num_decimal_places = 1

        if(digit < 0 or digit == None): digit = 0

        if(digit > 9): digit = 9

        gap_count = 0
        gaps = []


        for i in range(len(data_array)):
            digits_str = str(data_array[i]).replace('.', '')
            if(len(digits_str) < (num_decimal_places + 1)):
                current_digit = 0
            else:
                current_digit = int(digits_str[num_decimal_places])

            if(current_digit == digit):
                gaps.append(gap_count)
                gap_count = 0
            else:
                gap_count += 1


        gaps.append(gap_count)

        intervals = [i for i in range(max(gaps)+1)]

        observed_frequency = []
        expected_probability = []

        for n in intervals:
            observed_frequency.append(gaps.count(n))
            expected_probability.append(pow(0.9, n)*0.1)  

        if(spreadsheet):
            observed_probability = [n/sum(observed_frequency) for n in observed_frequency]

            merge_intervals_max = max(gaps)
            merged_intervals_str = []
            mod_difference_count = 0
            merged_intervals_observed = [] #probabilidade observada mergida (10 a 10)
            merged_intervals_expected = [] #probabilidade esperada mergida (10 a 10)

            while(merge_intervals_max % 10 != 0):
                merge_intervals_max += 1
                mod_difference_count += 1

            for i in range(int(merge_intervals_max/10)-1):
                merged_intervals_str.append(str(i*10) + ' - ' + str(i*10+9))
                merged_intervals_observed.append( sum((observed_probability[(i*10):(i*10+9+1)])) )
                merged_intervals_expected.append( sum((expected_probability[(i*10):(i*10+9+1)])) )  

            merged_intervals_str.append(str(merge_intervals_max - 10) + ' - ' + str(merge_intervals_max - mod_difference_count))
            merged_intervals_observed.append( sum(observed_probability[(merge_intervals_max - 10):(merge_intervals_max - mod_difference_count+1)]))
            merged_intervals_expected.append( sum(expected_probability[(merge_intervals_max - 10):(merge_intervals_max - mod_difference_count+1)]))

            wb = xl.Workbook()
            sheet = wb['Sheet']

            sheet.__setattr__('title', 'Intervalos - Digito ' + str(digit))

            sheet.append(['Intervalos', 'Probabilidade Observada', 'Probabilidade Esperada'])
            sheet.column_dimensions['B'].width = 21
            sheet.column_dimensions['C'].width = 19

            for i in range(len(merged_intervals_str)):
                sheet.cell(row = (i+2), column = 1, value=merged_intervals_str[i])
                sheet.cell(row = (i+2), column = 2, value=merged_intervals_observed[i])
                sheet.cell(row = (i+2), column = 3, value=merged_intervals_expected[i])

            chart1 = barChart.BarChart()
            chart1.type = "col"
            chart1.style = 10 
            chart1.title = 'Teste de Intervalos - Digito ' + str(digit)
            chart1.y_axis.title = 'Probabilidade'
            chart1.x_axis.title = 'Intervalos'

            graph_row = str(len(merged_intervals_str) + 5)

            data = reference.Reference(sheet, min_col=2, min_row=2, max_row=len(merged_intervals_str)+1, max_col=3)
            cats = reference.Reference(sheet, min_col=1, min_row=2, max_row=len(merged_intervals_str)+1)
            chart1.add_data(data)
            chart1.set_categories(cats)
            chart1.shape = 4
            chart1.series[0].title = series.SeriesLabel(v='Probabilidade Observada')
            chart1.series[1].title = series.SeriesLabel(v='Probabilidade Esperada')
            chart1.width = 30
            chart1.height = 15
            sheet.add_chart(chart1, 'A' + graph_row)

            wb.save(spreadsheet_title + '.xlsx')

                
        return self.__kolmogorov_smirnov(observed_frequency, expected_probability)


    def permutations_test(self, interval_size = 3, spreadsheet: bool = False, spreadsheet_title: str = 'permutacoes'):
        data_array = self.data.copy()

        def compute_permutations(data: list, interval_size):
            data2 = data
            while(len(data2) % interval_size != 0):
                data2.remove(data2[-1])

            intervals = numpy.split(numpy.array(data2), int(len(data2) / interval_size))
            
            p = itertools.permutations(intervals[0], interval_size)

            permutations_count = [0] * len(list(p))

            for i in range(len(intervals)):
                p = itertools.permutations(intervals[i], interval_size)
                function_index = 0
                for s in list(p):
                    if( all([s[j] < s[j+1] for j in range(len(s)-1)]) ):
                        permutations_count[function_index] = permutations_count[function_index] + 1
                        break

                    function_index = function_index + 1

            return permutations_count

        observed_frequency = compute_permutations(data_array, interval_size)
        expected_probability = [(1/math.factorial(interval_size))] * len(observed_frequency)
        observed_probability = [n / sum(observed_frequency) for n in observed_frequency]

        if(spreadsheet):
            wb = xl.Workbook()
            sheet = wb['Sheet']

            sheet.__setattr__('title', 'Permutações')

            funcoes = ['fO' + str(i) for i in range(1, len(observed_frequency)+1)]
            
            sheet.append(['classes', 'Probabilidade Observada', 'Probabilidade Esperada'])
            sheet.column_dimensions['B'].width = 21
            sheet.column_dimensions['C'].width = 19

            for i in range(len(funcoes)):
                sheet.cell(row = (i+2), column = 1, value=funcoes[i])
                sheet.cell(row = (i+2), column = 2, value=observed_probability[i])
                sheet.cell(row = (i+2), column = 3, value=expected_probability[i])

            chart1 = barChart.BarChart()
            chart1.type = "col"
            chart1.style = 10 
            chart1.title = 'Teste de Permutações'
            chart1.y_axis.title = 'Probabilidade'
            chart1.x_axis.title = 'Funções'

            graph_row = str(len(funcoes) + 5)

            data = reference.Reference(sheet, min_col=2, min_row=2, max_row=len(funcoes)+1, max_col=3)
            cats = reference.Reference(sheet, min_col=1, min_row=2, max_row=len(funcoes)+1)
            chart1.add_data(data)
            chart1.set_categories(cats)
            chart1.shape = 4
            chart1.series[0].title = series.SeriesLabel(v='Probabilidade Observada')
            chart1.series[1].title = series.SeriesLabel(v='Probabilidade Esperada')
            chart1.width = 30
            chart1.height = 15
            sheet.add_chart(chart1, 'A' + graph_row)

            wb.save(spreadsheet_title + '.xlsx')
        

        return self.__kolmogorov_smirnov(observed_frequency, expected_probability)

