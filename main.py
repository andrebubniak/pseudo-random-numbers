import time
from timeit import timeit
from typing import List
import randomnessTests
import pseudoRandomNumberGenerators as generators
import openpyxl as xl


def write_generator_data(array: List, file_name: str = 'GERALEO.txt'):
    try:
        with open(file_name, 'w') as temp_file:
            temp_file.write("\n".join(map(str,array)))
    except Exception as e: 
        print("Falha ao gravar no arquivo: \n", e)

def save_generator_time(qty: int = 10):
    wb = xl.Workbook()
    sheet = wb['Sheet']

    sheet.__setattr__('title', 'Tempo')
    
    sheet.append(['GLIM', 'Meu gerador'])

    for i in range(10):
        t1 = time.time()
        generators.glim3_generator(50000000)
        t2 = time.time()
        generators.generator_multiplicative_congruential(50000000)
        t3 = time.time()
        
        sheet.cell(row = (i+2), column = 1, value=(t2-t1))
        sheet.cell(row = (i+2), column = 2, value=(t3-t2))

    wb.save('tempo_geradores.xlsx')


def test_generator(glim: bool, qty: int = 50000000, save_data: bool = False):
    
    def bool2str(value: bool):
        if(value): return 'Aceitou'
        else: return 'Rejeitou'
    
    if(glim):
        excel_name = 'glim3'
        print('--- Teste do gerador GLIM3 ---')
        t1 = time.time()
        data = generators.glim3_generator(qty)
        t2 = time.time()
    else:
        excel_name = 'meuGerador'
        print('--- Teste do gerador "Inedito" ---')
        t1 = time.time()
        data = generators.generator_multiplicative_congruential(qty)
        t2 = time.time()

    if(save_data):
        write_generator_data(data)
        
    try:
        rt = randomnessTests.randomnessTests(data)
    except Exception as e:
        print(e)
        exit(-1)


    print("Quantidade Gerada: " + str(qty))
    print('Tempo utilizado: ' + str(t2-t1) + ' segundos')
    print('--- Testes ---')
    print('Teste de Uniformidade: ' + bool2str(rt.uniformity_test(2, True, excel_name + '_uniformidade')))
    print('Teste das corridas: ' + bool2str(rt.runs_test(True, excel_name + '_corridas')))
    print('Teste dos Intervalos: ')
    for i in range(10):
        if(i == 9): print('digito 9: ' + bool2str(rt.gap_test(1, i, True, excel_name + '_intervalos')))
        else: print('digito ' + str(i) + ': ' + bool2str(rt.gap_test(1, i)))

    print('Teste de permutacao: ' + bool2str(rt.permutations_test(3, True, excel_name + '_permutacoes')))



def main():
    test_generator(True, 1000000)
    test_generator(False, 50000000, True)
    save_generator_time()



if __name__ == '__main__':
    main()